"""Report generation service – PDF and Excel."""

import io
import os
import logging
from datetime import datetime
from typing import Dict

logger = logging.getLogger(__name__)


def generate_excel_report(kpis: Dict, sales: list, products: Dict, customers: Dict) -> bytes:
    """Generate Excel report with multiple sheets."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()

    # ── Sheet 1: KPI Summary ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "KPI Summary"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14, color="1E3A5F")

    ws1["A1"] = "Commerce Dashboard – Executive Summary"
    ws1["A1"].font = Font(bold=True, size=16, color="1E3A5F")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws1["A2"].font = Font(size=10, color="666666")

    ws1.append([])
    ws1.append(["Metric", "Value"])
    for cell in ws1[4]:
        cell.fill = header_fill
        cell.font = header_font

    kpi_rows = [
        ("Total Revenue", f"₹{kpis.get('total_revenue', 0):,.2f}"),
        ("Total Orders", f"{kpis.get('total_orders', 0):,}"),
        ("Total Customers", f"{kpis.get('total_customers', 0):,}"),
        ("Total Profit", f"₹{kpis.get('total_profit', 0):,.2f}"),
        ("Profit Margin", f"{kpis.get('profit_margin', 0):.1f}%"),
        ("Avg Order Value", f"₹{kpis.get('avg_order_value', 0):,.2f}"),
        ("Total Units Sold", f"{kpis.get('total_units', 0):,}"),
        ("Revenue Growth", f"{kpis.get('revenue_growth', 0):+.1f}%"),
    ]
    for row in kpi_rows:
        ws1.append(row)

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Sales Trend ─────────────────────────────────
    ws2 = wb.create_sheet("Sales Trend")
    if sales:
        headers = list(sales[0].keys())
        ws2.append(headers)
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in sales:
            ws2.append([row.get(h) for h in headers])
        for i, _ in enumerate(headers, 1):
            ws2.column_dimensions[get_column_letter(i)].width = 15

    # ── Sheet 3: Top Products ────────────────────────────────
    ws3 = wb.create_sheet("Top Products")
    top_prods = products.get("top_products", [])
    if top_prods:
        headers = ["product_name", "category", "revenue", "profit", "units_sold", "order_count"]
        ws3.append(headers)
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for p in top_prods:
            ws3.append([p.get(h) for h in headers])
        for i, _ in enumerate(headers, 1):
            ws3.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 4: Category Breakdown ──────────────────────────
    ws4 = wb.create_sheet("Category Breakdown")
    cats = products.get("category_breakdown", [])
    if cats:
        headers = list(cats[0].keys())
        ws4.append(headers)
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in cats:
            ws4.append([row.get(h) for h in headers])

    # ── Sheet 5: Top Customers ───────────────────────────────
    ws5 = wb.create_sheet("Top Customers")
    top_custs = customers.get("top_customers", [])
    if top_custs:
        headers = ["customer_name", "segment", "total_revenue", "total_profit", "order_count"]
        ws5.append(headers)
        for cell in ws5[1]:
            cell.fill = header_fill
            cell.font = header_font
        for c in top_custs:
            ws5.append([c.get(h) for h in headers])
        for i, _ in enumerate(headers, 1):
            ws5.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf_report(kpis: Dict, sales: list, products: Dict, insights: list) -> bytes:
    """Generate PDF report using reportlab."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        PRIMARY = colors.HexColor("#1E3A5F")
        ACCENT = colors.HexColor("#3B82F6")
        LIGHT = colors.HexColor("#F1F5F9")

        title_style = ParagraphStyle(
            "title", parent=styles["Title"],
            textColor=PRIMARY, fontSize=20, spaceAfter=4
        )
        subtitle_style = ParagraphStyle(
            "subtitle", parent=styles["Normal"],
            textColor=colors.grey, fontSize=10, spaceAfter=16
        )
        section_style = ParagraphStyle(
            "section", parent=styles["Heading2"],
            textColor=PRIMARY, fontSize=13, spaceBefore=16, spaceAfter=8
        )
        body_style = styles["BodyText"]

        story = []

        # Title
        story.append(Paragraph("Commerce Dashboard Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
        story.append(Spacer(1, 12))

        # KPIs table
        story.append(Paragraph("Key Performance Indicators", section_style))

        kpi_data = [["Metric", "Value", "Metric", "Value"]]
        kpi_items = [
            ("Total Revenue", f"₹{kpis.get('total_revenue', 0):,.0f}"),
            ("Total Profit", f"₹{kpis.get('total_profit', 0):,.0f}"),
            ("Total Orders", f"{kpis.get('total_orders', 0):,}"),
            ("Total Customers", f"{kpis.get('total_customers', 0):,}"),
            ("Profit Margin", f"{kpis.get('profit_margin', 0):.1f}%"),
            ("Avg Order Value", f"₹{kpis.get('avg_order_value', 0):,.0f}"),
            ("Revenue Growth", f"{kpis.get('revenue_growth', 0):+.1f}%"),
            ("Total Units", f"{kpis.get('total_units', 0):,}"),
        ]
        for i in range(0, len(kpi_items), 2):
            row = list(kpi_items[i]) + (list(kpi_items[i + 1]) if i + 1 < len(kpi_items) else ["", ""])
            kpi_data.append(row)

        kpi_table = Table(kpi_data, colWidths=[3.5 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BACKGROUND", (0, 1), (-1, -1), LIGHT),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),
            ("FONTNAME", (3, 1), (3, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (1, 1), (1, -1), ACCENT),
            ("TEXTCOLOR", (3, 1), (3, -1), ACCENT),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 12))

        # Top Products
        story.append(Paragraph("Top 10 Products by Revenue", section_style))
        top_prods = products.get("top_products", [])[:10]
        if top_prods:
            prod_data = [["Product", "Category", "Revenue (₹)", "Profit (₹)", "Units"]]
            for p in top_prods:
                prod_data.append([
                    p.get("product_name", "")[:30],
                    p.get("category", ""),
                    f"{p.get('revenue', 0):,.0f}",
                    f"{p.get('profit', 0):,.0f}",
                    str(p.get("units_sold", 0)),
                ])
            prod_table = Table(prod_data, colWidths=[4.5 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm])
            prod_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(prod_table)

        # Insights
        if insights:
            story.append(Spacer(1, 12))
            story.append(Paragraph("AI-Generated Business Insights", section_style))
            for ins in insights:
                story.append(Paragraph(f"• <b>{ins.get('title')}</b>: {ins.get('message')}", body_style))
                story.append(Spacer(1, 4))

        doc.build(story)
        buf.seek(0)
        return buf.read()

    except Exception as e:
        logger.exception("PDF generation failed")
        # Return minimal PDF on error
        return b""
